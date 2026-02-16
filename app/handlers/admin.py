import json
import re
import unicodedata
import tornado.web
from app.db import create_db_connection
from app.handlers.base import BaseHandler
from app.services import events_service, staff_service, telemetry_service


_HEX_COLOR_RE = re.compile(r"^#(?:[0-9a-fA-F]{3}|[0-9a-fA-F]{6})$")


def _sanitize_hex_color(value):
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    if not _HEX_COLOR_RE.match(s):
        return None
    # Normalize short hex to 6-char
    if len(s) == 4:
        return f"#{s[1]}{s[1]}{s[2]}{s[2]}{s[3]}{s[3]}"
    return s.lower()

def _slugify(value):
    """
    Normalizes string, converts to lowercase, removes non-alpha characters,
    and converts spaces to hyphens.
    """
    value = str(value)
    # Normalize unicode to ascii (e.g. á -> a)
    value = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')
    # Replace non-word chars (but keep hyphens/underscores initially to convert them later if needed)
    # Actually, user wants 'pollito_feliz' -> 'pollito-feliz', so let's treat underscore as delimiter or replaceable
    # Replace anything that is NOT a letter, number, or underscore with a hyphen (space included)
    value = re.sub(r'[^\w\s-]', '-', value)
    value = re.sub(r'[_\s]+', '-', value)
    # Remove duplicates
    value = re.sub(r'[-\s]+', '-', value)
    return value.strip('-').lower()




class EventsAdminHandler(BaseHandler):
    @tornado.web.authenticated
    def get(self):
        # Admin console: superadmin sees all events; event-admin sees only assigned events.
        if self.is_superadmin():
            events = events_service.list_events()

            # Enrich events with staff info and registration counts
            if events:
                with create_db_connection() as conn:
                    with conn.cursor() as cursor:
                        for evt in events:
                            eid = evt["id"]
                            # Get Staff Names
                            cursor.execute(
                                "SELECT u.name, es.role FROM event_staff es "
                                "JOIN users u ON u.id = es.user_id "
                                "WHERE es.event_id=%s AND es.role IN ('moderator', 'speaker')",
                                (eid,)
                            )
                            staff_rows = cursor.fetchall()
                            evt["moderator_name"] = next((r["name"] for r in staff_rows if r["role"] == "moderator"), None)
                            evt["speaker_name"] = next((r["name"] for r in staff_rows if r["role"] == "speaker"), None)
                            
                            # Get Registry Count (Viewers registered for this event)
                            cursor.execute(
                                "SELECT COUNT(*) as cnt FROM users WHERE event_id=%s AND role='viewer'",
                                (eid,)
                            )
                            count_res = cursor.fetchone()
                            evt["registration_count"] = count_res["cnt"] if count_res else 0

            self.render("admin/events.html", events=events, is_superadmin=True)
            return

        user_id = self.get_current_user()
        if not user_id:
            self.redirect("/watch")
            return

        from app.services import staff_service
        allowed_event_ids = staff_service.list_event_ids_for_role(int(user_id), "admin")
        
        # If they don't have assigned events but ARE admins (global role), 
        # let them see the dashboard (empty state message) instead of /watch.
        if not allowed_event_ids and not self.is_admin():
            self.redirect("/watch")
            return

        events = events_service.list_events(event_ids=allowed_event_ids) if allowed_event_ids else []
        
        # Enrich events with staff info and registration counts
        if events:
            with create_db_connection() as conn:
                with conn.cursor() as cursor:
                    for evt in events:
                        eid = evt["id"]
                        # Get Staff Names
                        cursor.execute(
                            "SELECT u.name, es.role FROM event_staff es "
                            "JOIN users u ON u.id = es.user_id "
                            "WHERE es.event_id=%s AND es.role IN ('moderator', 'speaker')",
                            (eid,)
                        )
                        staff_rows = cursor.fetchall()
                        evt["moderator_name"] = next((r["name"] for r in staff_rows if r["role"] == "moderator"), None)
                        evt["speaker_name"] = next((r["name"] for r in staff_rows if r["role"] == "speaker"), None)
                        
                        # Get Registry Count (Viewers registered for this event)
                        cursor.execute(
                            "SELECT COUNT(*) as cnt FROM users WHERE event_id=%s AND role='viewer'",
                            (eid,)
                        )
                        count_res = cursor.fetchone()
                        evt["registration_count"] = count_res["cnt"] if count_res else 0
        self.render("admin/events.html", events=events, is_superadmin=False)

class APIPollsHandler(BaseHandler):
    @tornado.web.authenticated
    def get(self):
        try:
            event_id = int(self.get_argument("event_id"))
            poll_id = self.get_argument("poll_id", None)
            poll_id = int(poll_id) if poll_id is not None else None
        except (TypeError, ValueError, tornado.web.MissingArgumentError):
            self.set_status(400)
            self.write({"status": "error", "message": "event_id requerido"})
            return

        if not (self.is_superadmin() or self.is_admin_for_event(event_id)):
            self.set_status(403)
            return

        from app.services import poll_service
        if poll_id:
            results = poll_service.get_poll_results(poll_id)
            if results is None:
                self.set_status(404)
                self.write({"status": "error", "message": "Resultados no disponibles"})
                return
            self.write(json.dumps({"status": "success", "results": results}, default=str))
            return

        polls = poll_service.list_polls(event_id)
        self.write(json.dumps({"status": "success", "polls": polls}, default=str))

    @tornado.web.authenticated
    def post(self):
        try:
            data = json.loads(self.request.body)
            event_id = int(data.get("event_id"))
            question = data.get("question", "").strip()
            options = data.get("options", [])
            status = data.get("status", "draft")
        except (TypeError, ValueError, json.JSONDecodeError):
            self.set_status(400)
            self.write({"status": "error", "message": "Datos inválidos"})
            return

        if not (self.is_superadmin() or self.is_admin_for_event(event_id)):
            self.set_status(403)
            return

        if not question or len(options) < 2:
            self.set_status(400)
            self.write({"status": "error", "message": "Pregunta y al menos 2 opciones requeridas"})
            return

        from app.services import poll_service
        poll_id = poll_service.create_poll(event_id, question, options, status)
        if poll_id:
            self.write({"status": "success", "poll_id": poll_id})
        else:
            self.set_status(500)
            self.write({"status": "error", "message": "Error al crear encuesta"})

    @tornado.web.authenticated
    def put(self):
        try:
            data = json.loads(self.request.body)
            poll_id = int(data.get("poll_id"))
            status = data.get("status")
            event_id = int(data.get("event_id")) # Required for permission check
            question = data.get("question", None)
            options = data.get("options", None)
        except (TypeError, ValueError, json.JSONDecodeError):
            self.set_status(400)
            self.write({"status": "error", "message": "Datos inválidos"})
            return

        if not (self.is_superadmin() or self.is_admin_for_event(event_id)):
            self.set_status(403)
            return
            
        from app.services import poll_service
        if question is not None or options is not None:
            poll = poll_service.get_poll_by_id(poll_id)
            if not poll:
                self.set_status(404)
                self.write({"status": "error", "message": "Encuesta no encontrada"})
                return
            if poll.get("status") != "draft":
                self.set_status(400)
                self.write({"status": "error", "message": "Solo puedes editar encuestas en borrador"})
                return
            if not question or not options or len(options) < 2:
                self.set_status(400)
                self.write({"status": "error", "message": "Pregunta y al menos 2 opciones requeridas"})
                return
            success = poll_service.update_poll_content(poll_id, question, options)
        else:
            if status not in ['draft', 'published', 'closed']:
                self.set_status(400)
                self.write({"status": "error", "message": "Estado inválido"})
                return
            success = poll_service.update_poll_status(poll_id, status)
        if success:
            self.write({"status": "success"})
        else:
            self.set_status(500)
            self.write({"status": "error", "message": "Error al actualizar estado"})

class APIEventsHandler(BaseHandler):
    @tornado.web.authenticated
    def post(self):
        # Only superadmin can create new events.
        if not self.is_superadmin():
            self.set_status(403)
            return

        try:
            data = json.loads(self.request.body)
            # Apply strict slugify
            raw_slug = data.get("slug")
            slug = _slugify(raw_slug) if raw_slug else None
            
            title = data.get("title")
            logo_url = data.get("logo_url")
            video_url = data.get("video_url")
            description = data.get("description")
            header_bg_color = _sanitize_hex_color(data.get("header_bg_color"))
            header_text_color = _sanitize_hex_color(data.get("header_text_color"))
            timezone = data.get("timezone", "America/Mexico_City")
            status = data.get("status")
            registration_mode = data.get("registration_mode")
            registration_restricted_type = data.get("registration_restricted_type")
            allowed_domain = data.get("allowed_domain")
            registration_open_at = data.get("registration_open_at")
            registration_close_at = data.get("registration_close_at")
            access_open_at = data.get("access_open_at")
            capacity = data.get("capacity")
            registration_schema = data.get("registration_schema")
            registration_success_message = data.get("registration_success_message")
            # Ensure registration_schema is a valid JSON string or None
            if registration_schema:
                try:
                    if isinstance(registration_schema, str):
                        # Validate it's proper JSON
                        parsed = json.loads(registration_schema)
                        registration_schema = json.dumps(parsed, ensure_ascii=False)
                    else:
                        registration_schema = json.dumps(registration_schema, ensure_ascii=False)
                except Exception:
                    registration_schema = None
            else:
                registration_schema = None
            
            is_deleted = data.get("is_deleted")

            if not slug or not title:
                self.set_status(400)
                self.write({"status": "error", "message": "Slug and Title are required"})
                return

            # Uniqueness Check
            existing = events_service.get_event_by_slug(slug)
            if existing:
                self.set_status(409) # Conflict
                self.write({"status": "error", "message": f"El slug '{slug}' ya está en uso. Elige otro."})
                return

            event_id = events_service.create_event(
                slug,
                title,
                logo_url,
                video_url,
                description,
                header_bg_color,
                header_text_color,
                timezone,
                status,
                registration_mode,
                registration_restricted_type,
                allowed_domain,
                registration_open_at,
                registration_close_at,
                access_open_at,
                capacity,
                registration_schema,
                registration_success_message,
                is_deleted,
            )
            self.write({"status": "success", "event_id": event_id})
        except Exception as e:
            self.set_status(500)
            self.write({"status": "error", "message": str(e)})

    @tornado.web.authenticated
    def put(self):
        try:
            data = json.loads(self.request.body)
            event_id = data.get("id")
            if not event_id:
                self.set_status(400)
                self.write({"status": "error", "message": "id requerido"})
                return

            # Superadmin can update any event; event-admin can update assigned events only.
            if not (self.is_superadmin() or self.is_admin_for_event(event_id)):
                self.set_status(403)
                return

            title = data.get("title")
            logo_url = data.get("logo_url")
            video_url = data.get("video_url")
            description = data.get("description")
            is_active = data.get("is_active")
            header_bg_color = _sanitize_hex_color(data.get("header_bg_color"))
            header_text_color = _sanitize_hex_color(data.get("header_text_color"))
            timezone = data.get("timezone", "America/Mexico_City")
            status = data.get("status")
            registration_mode = data.get("registration_mode")
            registration_restricted_type = data.get("registration_restricted_type")
            allowed_domain = data.get("allowed_domain")
            registration_open_at = data.get("registration_open_at")
            registration_close_at = data.get("registration_close_at")
            access_open_at = data.get("access_open_at")
            capacity = data.get("capacity")
            registration_schema = data.get("registration_schema")
            registration_success_message = data.get("registration_success_message")
            # Ensure registration_schema is a valid JSON string or None
            if registration_schema:
                try:
                    if isinstance(registration_schema, str):
                        # Validate it's proper JSON
                        parsed = json.loads(registration_schema)
                        registration_schema = json.dumps(parsed, ensure_ascii=False)
                    else:
                        registration_schema = json.dumps(registration_schema, ensure_ascii=False)
                except Exception:
                    registration_schema = None
            else:
                registration_schema = None
            
            is_deleted = data.get("is_deleted")

            events_service.update_event(
                event_id,
                title,
                logo_url,
                video_url,
                is_active,
                description,
                header_bg_color,
                header_text_color,
                timezone,
                status,
                registration_mode,
                registration_restricted_type,
                allowed_domain,
                registration_open_at,
                registration_close_at,
                access_open_at,
                capacity,
                registration_schema,
                registration_success_message,
                is_deleted,
            )

            # If the event was closed, kick all users
            if not is_active or status == "CLOSED":
                from app.handlers import ws
                ws.kick_all_from_event(event_id)

            self.write({"status": "success"})
        except Exception as e:
            self.set_status(500)
            self.write({"status": "error", "message": str(e)})


class APIEventStaffHandler(BaseHandler):
    """Superadmin-only API to manage per-event staff assignments."""

    @tornado.web.authenticated
    def get(self):
        if not self.is_superadmin():
            self.set_status(403)
            return

        mode = self.get_query_argument("mode", "list")
        if mode == "users":
            # Return list of users filtered by role and availability
            search_query = self.get_query_argument("q", "").strip().lower()
            target_role = self.get_query_argument("role", "admin").lower()

            with create_db_connection() as conn:
                with conn.cursor() as cursor:
                    # Base query parts
                    query = "SELECT u.id, u.email, u.name FROM users u "
                    params = []
                    
                    # 1. Filter by Global Role
                    where_clauses = ["u.role = %s"]
                    params.append(target_role)

                    # 2. Search filter (name/email)
                    if search_query:
                        where_clauses.append("(u.email LIKE %s OR u.name LIKE %s)")
                        params.append(f"%{search_query}%")
                        params.append(f"%{search_query}%")

                    # 3. Availability Check (Exclusive for Moderator/Speaker)
                    # Admins can be in multiple events, but Mods/Speakers must be free.
                    if target_role in ["moderator", "speaker"]:
                        where_clauses.append("u.id NOT IN (SELECT user_id FROM event_staff)")

                    # Construct final query
                    sql = f"{query} WHERE {' AND '.join(where_clauses)} ORDER BY u.name ASC LIMIT 20"
                    
                    cursor.execute(sql, tuple(params))
                    users = cursor.fetchall() or []
                    
                    # Map to Select2 format
                    results = []
                    for u in users:
                        email = u["email"]
                        name = u["name"] or ""
                        results.append({
                            "id": email, # We still use email as ID for the frontend logic
                            "text": f"{name} ({email})" if name else email
                        })
                    self.write({"status": "success", "results": results})
            return

        try:
            event_id = int(self.get_query_argument("event_id"))
        except Exception:
            self.set_status(400)
            self.write({"status": "error", "message": "event_id requerido"})
            return

        from app.services import staff_service
        rows = staff_service.list_staff_for_event(event_id)
        self.write({"status": "success", "staff": rows})

    @tornado.web.authenticated
    def post(self):
        if not self.is_superadmin():
            self.set_status(403)
            return

        try:
            data = json.loads(self.request.body)
            event_id = int(data.get("event_id"))
            email = data.get("email")
            role = data.get("role")

            from app.services import staff_service
            assignment = staff_service.upsert_staff_by_email(event_id=event_id, email=email, role=role)
            self.write({"status": "success", "assignment": assignment})
        except Exception as e:
            self.set_status(400)
            self.write({"status": "error", "message": str(e)})

    @tornado.web.authenticated
    def delete(self):
        if not self.is_superadmin():
            self.set_status(403)
            return

        try:
            data = json.loads(self.request.body) if self.request.body else {}
            event_id = int(data.get("event_id") or self.get_query_argument("event_id"))
            user_id = int(data.get("user_id") or self.get_query_argument("user_id"))
        except Exception:
            self.set_status(400)
            self.write({"status": "error", "message": "event_id y user_id requeridos"})
            return

        from app.services import staff_service
        ok = staff_service.remove_staff(user_id=user_id, event_id=event_id)
        self.write({"status": "success", "removed": bool(ok)})


class APIEventWhitelistHandler(BaseHandler):
    @tornado.web.authenticated
    def get(self):
        try:
            event_id = int(self.get_query_argument("event_id"))
        except Exception:
            self.set_status(400)
            self.write({"status": "error", "message": "event_id requerido"})
            return

        if not (self.is_superadmin() or self.is_admin_for_event(event_id)):
            self.set_status(403)
            return

        emails = events_service.get_whitelist(event_id)
        self.write({"status": "success", "emails": emails})

    @tornado.web.authenticated
    def put(self):
        try:
            data = json.loads(self.request.body)
            event_id = int(data.get("event_id"))
            emails = data.get("emails") or []
        except Exception:
            self.set_status(400)
            self.write({"status": "error", "message": "event_id y emails requeridos"})
            return

        if not (self.is_superadmin() or self.is_admin_for_event(event_id)):
            self.set_status(403)
            return

        count = events_service.replace_whitelist(event_id, emails)
        self.write({"status": "success", "count": count})

    @tornado.web.authenticated
    def post(self):
        """Procesar carga masiva de whitelist desde archivo (xlsx, xls, csv)"""
        try:
            event_id = int(self.get_argument("event_id"))
        except Exception:
            self.set_status(400)
            self.write({"status": "error", "message": "event_id requerido"})
            return

        if not (self.is_superadmin() or self.is_admin_for_event(event_id)):
            self.set_status(403)
            return

        # Verificar que se subió un archivo
        if not hasattr(self.request, 'files') or 'file' not in self.request.files:
            self.set_status(400)
            self.write({"status": "error", "message": "No se subió ningún archivo"})
            return

        file_info = self.request.files['file'][0]
        filename = file_info['filename']
        content = file_info['body']

        # Validar tipo de archivo
        allowed_extensions = ['.xlsx', '.xls', '.csv']
        file_ext = '.' + filename.split('.')[-1].lower() if '.' in filename else ''
        
        if file_ext not in allowed_extensions:
            self.set_status(400)
            self.write({"status": "error", "message": "Tipo de archivo no permitido. Use xlsx, xls o csv"})
            return

        # Validar tamaño (5MB máximo)
        max_size = 5 * 1024 * 1024
        if len(content) > max_size:
            self.set_status(400)
            self.write({"status": "error", "message": "Archivo demasiado grande. Máximo 5MB"})
            return

        try:
            emails = self._parse_whitelist_file(content, file_ext)
            
            if not emails:
                self.write({"status": "success", "emails": [], "message": "No se encontraron emails válidos en el archivo"})
                return

            # Guardar en la whitelist
            count = events_service.replace_whitelist(event_id, emails)
            self.write({
                "status": "success", 
                "emails": emails, 
                "count": count,
                "message": f"Se procesaron {len(emails)} emails correctamente"
            })

        except Exception as e:
            self.set_status(500)
            self.write({"status": "error", "message": f"Error procesando archivo: {str(e)}"})

    def _parse_whitelist_file(self, content, file_ext):
        """Parsear archivo y extraer emails"""
        emails = []
        
        try:
            if file_ext == '.csv':
                # Procesar CSV
                import csv
                import io
                
                content_str = content.decode('utf-8', errors='ignore')
                csv_file = io.StringIO(content_str)
                reader = csv.reader(csv_file)
                
                for row in reader:
                    for cell in row:
                        email = self._extract_email(cell)
                        if email:
                            emails.append(email)
                            
            elif file_ext in ['.xlsx', '.xls']:
                # Procesar Excel
                import openpyxl
                import io
                
                excel_file = io.BytesIO(content)
                
                if file_ext == '.xlsx':
                    workbook = openpyxl.load_workbook(excel_file)
                else:
                    # Para .xls necesitaríamos xlrd, pero podemos intentar con openpyxl
                    workbook = openpyxl.load_workbook(excel_file, read_only=True)
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for row in sheet.iter_rows(values_only=True):
                        for cell in row:
                            if cell:
                                email = self._extract_email(str(cell))
                                if email:
                                    emails.append(email)
                                    
        except Exception as e:
            raise Exception(f"Error leyendo archivo: {str(e)}")
        
        # Limpiar y deduplicar emails
        emails = list(set(email.lower().strip() for email in emails if email and '@' in email))
        emails.sort()
        
        return emails

    def _extract_email(self, text):
        """Extraer email de un texto"""
        import re
        
        # Buscar patrones de email
        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        matches = re.findall(email_pattern, text)
        
        if matches:
            return matches[0].lower().strip()
        
        # Si no hay patrón claro, verificar si el texto completo es un email
        text = text.strip()
        if '@' in text and '.' in text.split('@')[-1]:
            return text.lower()
        
        return None


class StaffAdminHandler(BaseHandler):
    @tornado.web.authenticated
    def get(self):
        if not self.is_superadmin():
            self.redirect("/watch")
            return
        
        staff_list = staff_service.list_all_staff_global()
        self.render("admin/staff.html", staff=staff_list, is_superadmin=True)


class APIStaffHandler(BaseHandler):
    @tornado.web.authenticated
    def post(self):
        if not self.is_superadmin():
            self.set_status(403)
            return

        try:
            data = json.loads(self.request.body)
            user_id = data.get("id")
            email = data.get("email", "").strip().lower()
            name = data.get("name", "").strip()
            role = data.get("role", "viewer")

            if not email:
                self.set_status(400)
                self.write({"status": "error", "message": "Email es requerido"})
                return

            with create_db_connection() as conn:
                with conn.cursor() as cursor:
                    if user_id:
                        # Update by ID
                        cursor.execute(
                            "UPDATE users SET email=%s, name=%s, role=%s WHERE id=%s",
                            (email, name, role, user_id)
                        )
                    else:
                        # Check if user already exists by email (for global users)
                        cursor.execute(
                            "SELECT id FROM users WHERE email=%s AND event_id IS NULL LIMIT 1",
                            (email,)
                        )
                        existing = cursor.fetchone()
                        if existing:
                            cursor.execute(
                                "UPDATE users SET name=%s, role=%s WHERE id=%s",
                                (name, role, existing["id"])
                            )
                        else:
                            cursor.execute(
                                "INSERT INTO users (email, name, role, event_id) VALUES (%s, %s, %s, NULL)",
                                (email, name or email.split("@")[0], role)
                            )
                    conn.commit()
            
            self.write({"status": "success"})
        except Exception as e:
            self.set_status(500)
            self.write({"status": "error", "message": str(e)})


class TelemetryAdminHandler(BaseHandler):
    @tornado.web.authenticated
    def get(self):
        if not self.is_superadmin():
            self.redirect("/watch")
            return

        # Get historical data (last 6 hours for the dashboard charts)
        history = telemetry_service.get_recent_history(hours=6)
        
        # Process history for the frontend
        # Convert timestamps from UTC (DB) to Mexico City time
        from zoneinfo import ZoneInfo
        mx_tz = ZoneInfo('America/Mexico_City')
        processed_history = []
        for h in history:
            # Parse metrics_json if it's a string (from MySQL JSON column)
            metrics_data = h["metrics_json"]
            if isinstance(metrics_data, str):
                metrics_data = json.loads(metrics_data)
            # Convert timestamp (assumed UTC) to MX time
            ts_local = h["timestamp"].astimezone(mx_tz)
            processed_history.append({
                "timestamp": ts_local.isoformat(),
                "metrics": metrics_data
            })

        errors = telemetry_service.get_recent_errors(limit=50)
        processed_errors = []
        for e in errors:
            ts_local = e["timestamp"].astimezone(mx_tz) if e.get("timestamp") else None
            processed_errors.append({
                "timestamp": ts_local.isoformat() if ts_local else None,
                "handler": e.get("handler"),
                "method": e.get("method"),
                "status": e.get("status"),
                "exception_type": e.get("exception_type"),
                "message": e.get("message"),
                "path": e.get("path")
            })

        self.render(
            "admin/telemetry.html",
            history_json=json.dumps(processed_history),
            errors_json=json.dumps(processed_errors),
            is_superadmin=True
        )


class TelemetryAdminDataHandler(BaseHandler):
    @tornado.web.authenticated
    def get(self):
        if not self.is_superadmin():
            self.set_status(403)
            self.write({"status": "forbidden"})
            return

        history = telemetry_service.get_recent_history(hours=6)

        from zoneinfo import ZoneInfo
        mx_tz = ZoneInfo('America/Mexico_City')
        processed_history = []
        for h in history:
            metrics_data = h["metrics_json"]
            if isinstance(metrics_data, str):
                metrics_data = json.loads(metrics_data)
            ts_local = h["timestamp"].astimezone(mx_tz)
            processed_history.append({
                "timestamp": ts_local.isoformat(),
                "metrics": metrics_data
            })

        errors = telemetry_service.get_recent_errors(limit=50)
        processed_errors = []
        for e in errors:
            ts_local = e["timestamp"].astimezone(mx_tz) if e.get("timestamp") else None
            processed_errors.append({
                "timestamp": ts_local.isoformat() if ts_local else None,
                "handler": e.get("handler"),
                "method": e.get("method"),
                "status": e.get("status"),
                "exception_type": e.get("exception_type"),
                "message": e.get("message"),
                "path": e.get("path")
            })

        self.write({
            "history": processed_history,
            "errors": processed_errors
        })


